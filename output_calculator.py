import openpyxl

def compare_columns(input_file, output_file):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
        
    # Add a new column header
    sheet.cell(row=1, column=3, value='Outcome')
    
    # Iterate through rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row, values_only=True), start=2):
        first_value = row[0]
        second_value = row[1]
        
        if first_value > second_value:
            sheet.cell(row=row_idx, column=3, value=1)
            print("added 1")
        elif second_value > first_value:
            sheet.cell(row=row_idx, column=3, value=0)
            print("added 0")
        else:
            sheet.cell(row=row_idx, column=3, value=2)
            print("added 2")             

    # Save the workbook
    workbook.save(output_file)

# Example usage:
input_file = 'player_battle_logs.xlsx'
output_file = 'player_battle_logs-output-deneme.xlsx'
compare_columns(input_file, output_file)